#!/usr/bin/env python3
"""
Report Generator
Generates QFD analysis reports, including a Markdown report and an Excel HOQ matrix.
"""

import json
import sys
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime


def generate_priority_recommendations(
    requirements: List[Dict],
    feature_scores: List[Dict],
    conflicts: List[Dict],
    gap_analysis: Dict
) -> Dict[str, Any]:
    """
    Generate priority recommendations.

    The scoring combines:
    - customer importance (weight)
    - competitive gap
    - technical linkage
    - conflict complexity
    """
    # Compute a priority score for each requirement.
    priority_scores = []
    
    disadvantages = {d["requirement_id"]: d for d in gap_analysis.get("disadvantages", [])}
    critical_gaps = {g["requirement_id"] for g in gap_analysis.get("summary", {}).get("critical_gaps", [])}
    
    # Collect requirements affected by conflicts.
    conflict_affected = set()
    for conf in conflicts:
        conflict_affected.update(conf.get("affected_requirements", []))
    
    for req in requirements:
        req_id = req["id"]
        importance = req.get("importance", 1) or 1
        
        # Base score = importance.
        score = importance * 2
        
        # Add points for competitive gaps.
        if req_id in critical_gaps:
            score += 3  # Critical gap
        elif req_id in disadvantages:
            score += 1.5  # Regular gap
        
        # Subtract points when the requirement is exposed to technical conflicts.
        if req_id in conflict_affected:
            score -= 0.5  # Conflict increases implementation difficulty
        
        priority_scores.append({
            "requirement_id": req_id,
            "requirement": req["description"],
            "importance": importance,
            "score": round(score, 1),
            "is_critical_gap": req_id in critical_gaps,
            "has_conflict": req_id in conflict_affected
        })
    
    # Sort.
    priority_scores.sort(key=lambda x: x["score"], reverse=True)
    
    # Add rank.
    for i, item in enumerate(priority_scores):
        item["rank"] = i + 1
    
    # Build the rationale text.
    top_priorities = []
    for item in priority_scores[:5]:
        rationale_parts = [f"Customer importance: {item['importance']}"]
        if item["is_critical_gap"]:
            rationale_parts.append("Critical competitive gap")
        if item["has_conflict"]:
            rationale_parts.append("(Note: involves technical conflicts)")
        
        top_priorities.append({
            "rank": item["rank"],
            "requirement_id": item["requirement_id"],
            "requirement": item["requirement"],
            "score": item["score"],
            "rationale": "; ".join(rationale_parts)
        })
    
    return {
        "top_priorities": top_priorities,
        "all_scores": priority_scores
    }


def find_integration_opportunities(
    requirements: List[Dict],
    features: List[Dict],
    correlation_matrix: Dict
) -> List[Dict[str, Any]]:
    """
    Find integration opportunities.

    When technical features are positively correlated, improving one may also
    improve another and satisfy multiple customer requirements at once.
    """
    opportunities = []
    correlations = correlation_matrix.get("correlations", [])
    
    # Find positively correlated feature pairs.
    positive_pairs = []
    for corr in correlations:
        if corr.get("correlation") in ["+", "++"]:
            positive_pairs.append({
                "feature_a": corr["feature_a"],
                "feature_b": corr["feature_b"],
                "strength": "strong" if corr["correlation"] == "++" else "moderate"
            })
    
    # Simple implementation: treat each positive pair as one integration opportunity.
    for i, pair in enumerate(positive_pairs):
        opportunities.append({
            "id": f"INTEG-{i+1:03d}",
            "features": [pair["feature_a"]["name"], pair["feature_b"]["name"]],
            "correlation_strength": pair["strength"],
            "description": f"Improving {pair['feature_a']['name']} may also improve {pair['feature_b']['name']}",
            "expected_benefit": "Single improvement addresses multiple requirements"
        })
    
    return opportunities


def generate_markdown_report(
    project_meta: Dict,
    requirements_data: Dict,
    relationship_matrix: Dict,
    correlation_matrix: Dict,
    conflicts_data: Dict,
    benchmark_data: Dict,
    gap_analysis: Dict,
    priorities: Dict
) -> str:
    """Generate the full Markdown report."""
    
    lines = []
    
    # Title
    lines.append(f"# QFD Analysis Report: {project_meta.get('project_name', 'Untitled')}")
    lines.append(f"\nGenerated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"\nProject ID: {project_meta.get('project_id', 'N/A')}")
    lines.append("")
    
    # Table of contents
    lines.append("## Table of Contents")
    lines.append("1. Executive Summary")
    lines.append("2. Customer Requirements")
    lines.append("3. Technical Features")
    lines.append("4. Relationship Matrix")
    lines.append("5. Technical Correlations & Conflicts")
    lines.append("6. Competitive Benchmark")
    lines.append("7. Priority Recommendations")
    lines.append("")
    
    # 1. Executive Summary
    lines.append("## 1. Executive Summary")
    lines.append("")
    
    req_count = len(requirements_data.get("requirements", []))
    feature_count = len(relationship_matrix.get("feature_scores", []))
    conflict_count = len(conflicts_data.get("conflicts", []))
    
    lines.append(f"- **Total Requirements Analyzed**: {req_count}")
    lines.append(f"- **Technical Features Evaluated**: {feature_count}")
    lines.append(f"- **Technical Conflicts Identified**: {conflict_count}")
    
    gap_summary = gap_analysis.get("summary", {})
    lines.append(f"- **Competitive Advantages**: {gap_summary.get('advantage_count', 0)}")
    lines.append(f"- **Areas Needing Improvement**: {gap_summary.get('disadvantage_count', 0)}")
    lines.append("")
    
    # Top 3 priorities
    top_3 = priorities.get("top_priorities", [])[:3]
    if top_3:
        lines.append("### Top 3 Priorities")
        for p in top_3:
            lines.append(f"{p['rank']}. **{p['requirement']}** (Score: {p['score']})")
        lines.append("")
    
    # 2. Customer Requirements
    lines.append("## 2. Customer Requirements")
    lines.append("")
    lines.append("| ID | Requirement | Importance | Frequency | Category |")
    lines.append("|----|-----------:|:----------:|:---------:|----------|")
    
    for req in requirements_data.get("requirements", []):
        lines.append(f"| {req['id']} | {req['description']} | {req.get('importance', '-')} | {req.get('frequency_in_voc', '-')} | {req.get('category', '-')} |")
    lines.append("")
    
    # 3. Feature Importance
    lines.append("## 3. Technical Features (by Importance)")
    lines.append("")
    lines.append("| Rank | Feature | Score |")
    lines.append("|:----:|---------|------:|")
    
    for fs in relationship_matrix.get("feature_scores", [])[:10]:
        lines.append(f"| {fs['rank']} | {fs['feature']} | {fs['score']} |")
    lines.append("")
    
    # 4. Relationship Matrix (simplified view)
    lines.append("## 4. Relationship Matrix")
    lines.append("")
    lines.append("*Full matrix available in hoq_matrix.xlsx*")
    lines.append("")
    lines.append("Legend: ◎=Strong(9), ○=Medium(3), △=Weak(1)")
    lines.append("")
    
    # 5. Conflicts
    lines.append("## 5. Technical Correlations & Conflicts")
    lines.append("")
    
    conflicts = conflicts_data.get("conflicts", [])
    if conflicts:
        lines.append(f"**{len(conflicts)} conflicts identified:**")
        lines.append("")
        for conf in conflicts:
            severity = "HIGH" if conf["severity"] == "high" else "MEDIUM"
            lines.append(f"- [{severity}] **{conf['feature_a']['name']}** vs **{conf['feature_b']['name']}**")
            lines.append(f"  - {conf['description']}")
            lines.append(f"  - Affects: {', '.join(conf['affected_requirements'])}")
        lines.append("")
    else:
        lines.append("No significant technical conflicts identified.")
        lines.append("")
    
    # 6. Competitive Benchmark
    lines.append("## 6. Competitive Benchmark")
    lines.append("")
    
    products = benchmark_data.get("products", [])
    benchmark_state = benchmark_data.get("benchmark_state", "completed" if len(products) > 1 else "skipped")
    if benchmark_state == "completed" and len(products) > 1:
        # Header row
        header = "| Requirement | Importance |"
        for p in products:
            name = "Ours" if p == "our_product" else p[:8]
            header += f" {name} |"
        header += " Gap |"
        lines.append(header)
        
        sep = "|-------------|:----------:|"
        for _ in products:
            sep += ":------:|"
        sep += ":----:|"
        lines.append(sep)
        
        for row in benchmark_data.get("requirement_comparison", []):
            line = f"| {row['requirement'][:20]} | {row['importance']} |"
            for p in products:
                rating = row.get("ratings", {}).get(p, {})
                level = rating.get("level", "?")
                line += f" {level} |"
            gap = row.get("gap_analysis", "")
            gap_icon = "+" if "Ahead" in gap else ("-" if "Behind" in gap else "=")
            line += f" {gap_icon} |"
            lines.append(line)
        lines.append("")
    else:
        if benchmark_state == "skipped":
            lines.append("*Benchmark skipped: no competitor specs provided.*")
        else:
            lines.append("*No competitor data available*")
        lines.append("")
    
    # 7. Priority Recommendations
    lines.append("## 7. Priority Recommendations")
    lines.append("")
    lines.append("### Top Priorities")
    lines.append("")
    
    for p in priorities.get("top_priorities", []):
        lines.append(f"**{p['rank']}. {p['requirement']}**")
        lines.append(f"- Priority Score: {p['score']}")
        lines.append(f"- Rationale: {p['rationale']}")
        lines.append("")
    
    # Closing note
    lines.append("---")
    lines.append(f"*Report generated by QFD Analysis Skill*")
    
    return "\n".join(lines)


def generate_hoq_excel(
    requirements: List[Dict],
    features: List[Dict],
    relationship_matrix: Dict,
    correlation_matrix: Dict,
    output_path: str
):
    """
    Generate the HOQ Excel file.
    Requires the openpyxl package.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        return False
    
    wb = Workbook()
    ws = wb.active
    ws.title = "House of Quality"
    
    # Style definitions
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    strong_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    medium_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    weak_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header row
    ws.cell(row=1, column=1, value="Requirement").fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=2, value="Weight").fill = header_fill
    ws.cell(row=1, column=2).font = header_font
    
    feature_scores = relationship_matrix.get("feature_scores", [])
    for i, fs in enumerate(feature_scores):
        cell = ws.cell(row=1, column=3+i, value=fs["feature"])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # Requirement rows
    matrix = relationship_matrix.get("matrix", [])
    for row_idx, req_row in enumerate(matrix):
        r = row_idx + 2
        ws.cell(row=r, column=1, value=req_row["requirement"])
        ws.cell(row=r, column=2, value=req_row["importance"]).alignment = center_align
        
        rel_dict = {rel["feature"]: rel for rel in req_row["relations"]}
        
        for col_idx, fs in enumerate(feature_scores):
            c = col_idx + 3
            rel = rel_dict.get(fs["feature"])
            cell = ws.cell(row=r, column=c)
            cell.alignment = center_align
            cell.border = thin_border
            
            if rel:
                cell.value = rel["strength"]
                if rel["strength"] == 9:
                    cell.fill = strong_fill
                    cell.font = Font(color="FFFFFF", bold=True)
                elif rel["strength"] == 3:
                    cell.fill = medium_fill
                elif rel["strength"] == 1:
                    cell.fill = weak_fill
    
    # Score row
    score_row = len(matrix) + 2
    ws.cell(row=score_row, column=1, value="Score").font = Font(bold=True)
    ws.cell(row=score_row, column=2, value="-")
    
    for col_idx, fs in enumerate(feature_scores):
        cell = ws.cell(row=score_row, column=3+col_idx, value=fs["score"])
        cell.alignment = center_align
        cell.font = Font(bold=True)
    
    # Rank row
    rank_row = score_row + 1
    ws.cell(row=rank_row, column=1, value="Rank").font = Font(bold=True)
    ws.cell(row=rank_row, column=2, value="-")
    
    for col_idx, fs in enumerate(feature_scores):
        cell = ws.cell(row=rank_row, column=3+col_idx, value=fs["rank"])
        cell.alignment = center_align
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 8
    for i in range(len(feature_scores)):
        col_letter = chr(ord('C') + i)
        ws.column_dimensions[col_letter].width = 12
    
    wb.save(output_path)
    return True


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python report_generator.py <workspace_path>")
        print("Example: python report_generator.py ./qfd-analysis")
        sys.exit(1)

    workspace = Path(sys.argv[1])

    if not workspace.exists():
        print(f"Error: Workspace not found: {workspace}")
        sys.exit(1)

    # Load all required data files
    def load_json(path):
        if path.exists():
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}

    print(f"Loading data from {workspace}...")

    # Load data files
    requirements_data = load_json(workspace / "01_voc" / "requirements.json")
    our_spec = load_json(workspace / "02_technical" / "our_product_spec.json")
    relationship_matrix = load_json(workspace / "03_analysis" / "relationship_matrix.json")
    correlation_matrix = load_json(workspace / "03_analysis" / "correlation_matrix.json")
    conflicts_data = load_json(workspace / "03_analysis" / "conflicts.json")
    benchmark_data = load_json(workspace / "04_benchmark" / "benchmark_table.json")
    gap_analysis = load_json(workspace / "04_benchmark" / "gap_analysis.json")

    # Generate priorities
    requirements = requirements_data.get("requirements", [])
    feature_scores = relationship_matrix.get("feature_scores", [])
    conflicts = conflicts_data.get("conflicts", [])

    priorities = generate_priority_recommendations(
        requirements, feature_scores, conflicts, gap_analysis
    )

    # Project metadata
    project_meta = {
        "project_name": workspace.name.replace("_QFD", ""),
        "project_id": workspace.name
    }

    # Generate markdown report
    report = generate_markdown_report(
        project_meta,
        requirements_data,
        relationship_matrix,
        correlation_matrix,
        conflicts_data,
        benchmark_data,
        gap_analysis,
        priorities
    )

    # Save report
    output_dir = workspace / "05_output"
    output_dir.mkdir(parents=True, exist_ok=True)

    report_path = output_dir / "qfd_report.md"
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report)
    print(f"✓ Report saved to: {report_path}")

    # Try to generate Excel HOQ
    hoq_path = output_dir / "hoq_matrix.xlsx"
    features = our_spec.get("features", [])
    if generate_hoq_excel(requirements, features, relationship_matrix, correlation_matrix, str(hoq_path)):
        print(f"✓ Excel HOQ saved to: {hoq_path}")
    else:
        print("⚠ Excel generation skipped (openpyxl not installed)")

    print("\nDone!")
